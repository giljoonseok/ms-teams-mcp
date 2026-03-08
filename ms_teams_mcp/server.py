"""
Microsoft Teams & Outlook MCP Server
Microsoft Graph API를 통해 Teams 채팅/채널 및 Outlook 메일을 관리하는 MCP 서버
"""

import sys
import os
import re
import json
import io
if sys.stdout is not None:
    sys.stdout.reconfigure(encoding="utf-8")
if sys.stderr is not None:
    sys.stderr.reconfigure(encoding="utf-8")
import msal
import requests
from importlib.metadata import version as _pkg_version
from fastmcp import FastMCP

__version__ = _pkg_version("ms-teams-mcp")

# ─────────────────────────────────────────
# 설정 (lazy 초기화 — import 시 크래시 방지)
# ─────────────────────────────────────────
SCOPES = [
    "Mail.Read", "Mail.Send", "User.Read",
    "Chat.Read", "Chat.ReadWrite",
    "Channel.ReadBasic.All",
    "ChannelMessage.Read.All", "ChannelMessage.Send",
    "Team.ReadBasic.All",
    "Files.Read.All",
]
TOKEN_CACHE_FILE = os.path.expanduser("~/.ms_mcp_token.json")

_cache = None
_app = None
_pub_app = None

def _get_config():
    client_id = os.environ.get("MS_CLIENT_ID")
    client_secret = os.environ.get("MS_CLIENT_SECRET")
    tenant_id = os.environ.get("MS_TENANT_ID")
    if not client_id or not client_secret or not tenant_id:
        raise RuntimeError(
            "환경변수가 설정되지 않았습니다: MS_CLIENT_ID, MS_CLIENT_SECRET, MS_TENANT_ID\n"
            "MCP 클라이언트 설정에서 env를 확인하거나, CLI 사용 시 환경변수를 설정하세요."
        )
    return client_id, client_secret, tenant_id

def _get_cache():
    global _cache
    if _cache is None:
        _cache = msal.SerializableTokenCache()
        if os.path.exists(TOKEN_CACHE_FILE):
            with open(TOKEN_CACHE_FILE, "r") as f:
                _cache.deserialize(f.read())
    return _cache

def _get_app():
    global _app
    if _app is None:
        client_id, client_secret, tenant_id = _get_config()
        _app = msal.ConfidentialClientApplication(
            client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
            client_credential=client_secret,
            token_cache=_get_cache(),
        )
    return _app

def _get_pub_app():
    global _pub_app
    if _pub_app is None:
        client_id, _, tenant_id = _get_config()
        _pub_app = msal.PublicClientApplication(
            client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
            token_cache=_get_cache(),
        )
    return _pub_app

def save_cache():
    cache = _get_cache()
    if cache.has_state_changed:
        with open(TOKEN_CACHE_FILE, "w") as f:
            f.write(cache.serialize())

# ─────────────────────────────────────────
# 토큰 획득 (silent 갱신 → 실패 시 에러)
# ─────────────────────────────────────────
def _reload_cache():
    """디스크에서 토큰 캐시를 다시 읽어 CLI 인증 결과를 반영"""
    cache = _get_cache()
    if os.path.exists(TOKEN_CACHE_FILE):
        with open(TOKEN_CACHE_FILE, "r") as f:
            cache.deserialize(f.read())

def get_token():
    app = _get_app()
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
        if result and "access_token" in result:
            save_cache()
            return result["access_token"]
    # 캐시 재로드 후 재시도 (CLI에서 인증한 경우)
    _reload_cache()
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
        if result and "access_token" in result:
            save_cache()
            return result["access_token"]
    raise Exception(
        "인증이 필요합니다. 'ms-teams-mcp auth'로 인증하거나 "
        "MCP에서 authenticate 도구를 호출하세요."
    )

def _headers():
    token = get_token()
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def _check_response(res: requests.Response):
    """Graph API 응답 상태 확인 및 사용자 친화적 에러 반환"""
    if res.ok:
        return
    status = res.status_code
    try:
        error_data = res.json()
        error_msg = error_data.get("error", {}).get("message", res.text[:200])
    except Exception:
        error_msg = res.text[:200]

    error_map = {
        401: f"인증 오류 (401): 토큰이 만료되었거나 유효하지 않습니다. 재인증이 필요합니다.\n상세: {error_msg}",
        403: f"권한 부족 (403): 이 작업에 필요한 권한이 없습니다. Azure AD 앱 권한을 확인하세요.\n상세: {error_msg}",
        404: f"리소스를 찾을 수 없습니다 (404): 요청한 항목이 존재하지 않거나 접근 권한이 없습니다.\n상세: {error_msg}",
        429: f"요청 한도 초과 (429): 잠시 후 다시 시도해주세요.\n상세: {error_msg}",
    }

    raise Exception(error_map.get(status, f"Graph API 오류 ({status}): {error_msg}"))

def graph_get(path: str, params: dict = None, url: str = None):
    if url is None:
        url = f"https://graph.microsoft.com/v1.0{path}"
    res = requests.get(url, headers=_headers(), params=params)
    _check_response(res)
    return res.json()

def graph_post(path: str, body: dict):
    url = f"https://graph.microsoft.com/v1.0{path}"
    res = requests.post(url, headers=_headers(), json=body)
    _check_response(res)
    return res.json()

def graph_post_action(path: str, body: dict):
    """Graph API POST (202 No Content 응답용 — sendMail, reply, forward 등)"""
    url = f"https://graph.microsoft.com/v1.0{path}"
    res = requests.post(url, headers=_headers(), json=body)
    _check_response(res)

def _parse_recipients(addresses: str) -> list[dict]:
    """쉼표 구분 이메일 주소를 Graph API recipients 형식으로 변환"""
    return [{"emailAddress": {"address": a.strip()}} for a in addresses.split(",") if a.strip()]

def _pagination_footer(data: dict, skip: int, top: int) -> str:
    """다음 페이지 존재 시 안내 문구 반환"""
    next_link = data.get("@odata.nextLink", "")
    if next_link:
        return (
            f"\n\n--- 추가 데이터 있음 ---"
            f"\n다음 페이지: skip={skip + top} 또는 next_link 값 사용"
            f"\nnext_link: {next_link}"
        )
    return ""

def strip_html(html: str) -> str:
    if not html:
        return ""
    text = re.sub(r"<[^>]+>", "", html).strip()
    return re.sub(r"\n{3,}", "\n\n", text)

# ─────────────────────────────────────────
# MCP 서버
# ─────────────────────────────────────────
mcp = FastMCP("Microsoft Teams MCP")

# ═══════════════════════════════════════════
# Teams - 팀/채널
# ═══════════════════════════════════════════

@mcp.tool()
def list_teams() -> str:
    """내가 참여한 Teams 팀 목록 조회"""
    data = graph_get("/me/joinedTeams")
    teams = data.get("value", [])
    if not teams:
        return "참여한 팀이 없습니다."
    result = []
    for t in teams:
        result.append(f"- {t['displayName']} (설명: {t.get('description','')}) | ID: {t['id']}")
    return "\n".join(result)

@mcp.tool()
def list_channels(team_id: str) -> str:
    """
    특정 팀의 채널 목록 조회
    - team_id: list_teams에서 확인한 팀 ID
    """
    data = graph_get(f"/teams/{team_id}/channels")
    channels = data.get("value", [])
    if not channels:
        return "채널이 없습니다."
    result = []
    for c in channels:
        result.append(f"- #{c['displayName']} (유형: {c.get('membershipType','')}) | ID: {c['id']}")
    return "\n".join(result)

@mcp.tool()
def list_channel_messages(team_id: str, channel_id: str, top: int = 20, skip: int = 0, next_link: str = "") -> str:
    """
    팀 채널의 메시지 목록 조회
    - team_id: 팀 ID
    - channel_id: 채널 ID
    - top: 가져올 메시지 수 (최대 50)
    - skip: 건너뛸 메시지 수 (페이지네이션, 기본값 0)
    - next_link: 다음 페이지 링크 (이전 결과에서 제공, 직접 입력 불필요)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 50)
        params = {"$top": top}
        if skip > 0:
            params["$skip"] = skip
        data = graph_get(
            f"/teams/{team_id}/channels/{channel_id}/messages",
            params=params
        )
    messages = data.get("value", [])
    if not messages:
        return "메시지가 없습니다."
    result = []
    for i, m in enumerate(messages, 1):
        sender_from = m.get("from") or {}
        sender_user = sender_from.get("user") or {}
        sender = sender_user.get("displayName") or "알 수 없음"
        body = strip_html(m.get("body", {}).get("content", ""))[:200]
        created = m.get("createdDateTime", "")[:19]
        result.append(
            f"{i}. [{created}] {sender}\n"
            f"   {body}\n"
            f"   ID: {m['id']}"
        )
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def send_channel_message(team_id: str, channel_id: str, message: str) -> str:
    """
    팀 채널에 메시지 보내기
    - team_id: 팀 ID
    - channel_id: 채널 ID
    - message: 보낼 메시지 내용
    """
    body = {"body": {"content": message}}
    data = graph_post(f"/teams/{team_id}/channels/{channel_id}/messages", body)
    return f"메시지 전송 완료 (ID: {data.get('id', '')})"

# ═══════════════════════════════════════════
# Teams - 1:1/그룹 채팅
# ═══════════════════════════════════════════

@mcp.tool()
def list_chats(top: int = 20, skip: int = 0, next_link: str = "") -> str:
    """
    내 1:1/그룹 채팅 목록 조회
    - top: 가져올 채팅 수 (최대 50)
    - skip: 건너뛸 채팅 수 (페이지네이션, 기본값 0)
    - next_link: 다음 페이지 링크 (이전 결과에서 제공, 직접 입력 불필요)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 50)
        params = {"$top": top, "$expand": "members", "$orderby": "lastMessagePreview/createdDateTime desc"}
        if skip > 0:
            params["$skip"] = skip
        data = graph_get("/me/chats", params=params)
    chats = data.get("value", [])
    if not chats:
        return "채팅이 없습니다."
    result = []
    for i, c in enumerate(chats, 1):
        chat_type = c.get("chatType", "")
        topic = c.get("topic") or ""
        members = [m.get("displayName") or "" for m in c.get("members", [])]
        members_str = ", ".join(members[:5])
        preview = c.get("lastMessagePreview", {})
        preview_body = strip_html(preview.get("body", {}).get("content", ""))[:80] if preview else ""
        preview_time = preview.get("createdDateTime", "")[:19] if preview else ""
        label = topic if topic else members_str
        result.append(
            f"{i}. [{chat_type}] {label}\n"
            f"   최근: [{preview_time}] {preview_body}\n"
            f"   ID: {c['id']}"
        )
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def list_chat_messages(chat_id: str, top: int = 20, skip: int = 0, next_link: str = "") -> str:
    """
    특정 채팅의 메시지 목록 조회
    - chat_id: list_chats에서 확인한 채팅 ID
    - top: 가져올 메시지 수 (최대 50)
    - skip: 건너뛸 메시지 수 (페이지네이션, 기본값 0)
    - next_link: 다음 페이지 링크 (이전 결과에서 제공, 직접 입력 불필요)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 50)
        params = {"$top": top}
        if skip > 0:
            params["$skip"] = skip
        data = graph_get(f"/me/chats/{chat_id}/messages", params=params)
    messages = data.get("value", [])
    if not messages:
        return "메시지가 없습니다."
    result = []
    for i, m in enumerate(messages, 1):
        sender = m.get("from", {})
        sender_name = "시스템"
        if sender and sender.get("user"):
            sender_name = sender["user"].get("displayName", "알 수 없음")
        body = strip_html(m.get("body", {}).get("content", ""))[:300]
        created = m.get("createdDateTime", "")[:19]
        msg_type = m.get("messageType", "")
        if msg_type == "systemEventMessage":
            continue
        result.append(
            f"{i}. [{created}] {sender_name}\n"
            f"   {body}\n"
            f"   ID: {m['id']}"
        )
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def send_chat_message(chat_id: str, message: str) -> str:
    """
    1:1/그룹 채팅에 메시지 보내기
    - chat_id: 채팅 ID
    - message: 보낼 메시지 내용
    """
    body = {"body": {"content": message}}
    data = graph_post(f"/me/chats/{chat_id}/messages", body)
    return f"메시지 전송 완료 (ID: {data.get('id', '')})"

@mcp.tool()
def create_chat(members: str, message: str = "", topic: str = "") -> str:
    """
    새 1:1 또는 그룹 채팅 생성
    - members: 참여자 이메일 주소 (쉼표로 구분, 본인 제외)
    - message: 첫 메시지 (선택)
    - topic: 그룹 채팅 주제 (선택, 3명 이상일 때 권장)
    """
    member_list = [m.strip() for m in members.split(",") if m.strip()]
    if not member_list:
        return "참여자 이메일을 1개 이상 입력해주세요."
    chat_type = "oneOnOne" if len(member_list) == 1 else "group"
    me = graph_get("/me", params={"$select": "id"})
    body = {
        "chatType": chat_type,
        "members": [
            {
                "@odata.type": "#microsoft.graph.aadUserConversationMember",
                "roles": ["owner"],
                "user@odata.bind": f"https://graph.microsoft.com/v1.0/users('{me['id']}')"
            }
        ] + [
            {
                "@odata.type": "#microsoft.graph.aadUserConversationMember",
                "roles": ["owner"],
                "user@odata.bind": f"https://graph.microsoft.com/v1.0/users('{email}')"
            }
            for email in member_list
        ]
    }
    if topic and chat_type == "group":
        body["topic"] = topic
    data = graph_post("/chats", body)
    if message:
        graph_post(f"/chats/{data['id']}/messages", {"body": {"content": message}})
    result = f"채팅 생성 완료\n- ID: {data.get('id', '')}\n- 유형: {chat_type}\n- 참여자: {', '.join(member_list)}"
    if topic and chat_type == "group":
        result += f"\n- 주제: {topic}"
    if message:
        result += "\n- 첫 메시지 전송 완료"
    return result

# ═══════════════════════════════════════════
# Outlook 메일
# ═══════════════════════════════════════════

@mcp.tool()
def list_emails(folder: str = "inbox", top: int = 10, skip: int = 0, next_link: str = "") -> str:
    """
    Outlook 메일 목록 조회
    - folder: inbox / sentItems / drafts / deleteditems
    - top: 가져올 메일 수 (최대 1000)
    - skip: 건너뛸 메일 수 (페이지네이션, 기본값 0)
    - next_link: 다음 페이지 링크 (이전 결과에서 제공, 직접 입력 불필요)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 1000)
        params = {
            "$top": top,
            "$select": "id,subject,from,receivedDateTime,isRead,bodyPreview",
            "$orderby": "receivedDateTime desc"
        }
        if skip > 0:
            params["$skip"] = skip
        data = graph_get(f"/me/mailFolders/{folder}/messages", params=params)
    emails = data.get("value", [])
    if not emails:
        return "메일이 없습니다."
    result = []
    for i, e in enumerate(emails, 1):
        read_mark = "N" if not e.get("isRead") else " "
        sender = e.get("from", {}).get("emailAddress", {})
        result.append(
            f"{i}. [{read_mark}] [{e['receivedDateTime'][:10]}] {e['subject']}\n"
            f"   발신: {sender.get('name','')} <{sender.get('address','')}>\n"
            f"   ID: {e['id']}\n"
            f"   미리보기: {e.get('bodyPreview','')[:80]}"
        )
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def read_email(message_id: str) -> str:
    """
    특정 메일 전체 본문 읽기
    - message_id: list_emails에서 확인한 ID
    """
    data = graph_get(f"/me/messages/{message_id}", params={"$select": "subject,from,toRecipients,receivedDateTime,body,isRead"})
    sender = data.get("from", {}).get("emailAddress", {})
    to_list = [r["emailAddress"]["address"] for r in data.get("toRecipients", [])]
    body_text = strip_html(data.get("body", {}).get("content", ""))
    return (
        f"제목: {data.get('subject')}\n"
        f"발신: {sender.get('name')} <{sender.get('address')}>\n"
        f"수신: {', '.join(to_list)}\n"
        f"날짜: {data.get('receivedDateTime','')[:19]}\n"
        f"{'─'*50}\n"
        f"{body_text[:3000]}"
    )

@mcp.tool()
def search_emails(query: str, top: int = 10, skip: int = 0, next_link: str = "") -> str:
    """
    메일 검색
    - query: 검색어 (제목, 본문, 발신자 등)
    - top: 결과 수 (최대 1000)
    - skip: 건너뛸 결과 수 (페이지네이션, 기본값 0)
    - next_link: 다음 페이지 링크 (이전 결과에서 제공, 직접 입력 불필요)
    """
    if next_link:
        data = graph_get("", url=next_link)
    else:
        top = min(top, 1000)
        params = {
            "$search": f'"{query}"',
            "$top": top,
            "$select": "id,subject,from,receivedDateTime,bodyPreview"
        }
        if skip > 0:
            params["$skip"] = skip
        data = graph_get("/me/messages", params=params)
    emails = data.get("value", [])
    if not emails:
        return f"'{query}' 검색 결과가 없습니다."
    result = []
    for i, e in enumerate(emails, 1):
        sender = e.get("from", {}).get("emailAddress", {})
        result.append(
            f"{i}. [{e['receivedDateTime'][:10]}] {e['subject']}\n"
            f"   발신: {sender.get('name')} <{sender.get('address')}>\n"
            f"   ID: {e['id']}"
        )
    return "\n\n".join(result) + _pagination_footer(data, skip, top)

@mcp.tool()
def send_email(to: str, subject: str, body: str, cc: str = "", bcc: str = "") -> str:
    """
    메일 발송
    - to: 수신자 이메일 (쉼표로 여러 명 가능)
    - subject: 제목
    - body: 본문 (텍스트)
    - cc: 참조 (쉼표로 여러 명 가능, 선택)
    - bcc: 숨은참조 (쉼표로 여러 명 가능, 선택)
    """
    message = {
        "subject": subject,
        "body": {"contentType": "Text", "content": body},
        "toRecipients": _parse_recipients(to),
    }
    if cc:
        message["ccRecipients"] = _parse_recipients(cc)
    if bcc:
        message["bccRecipients"] = _parse_recipients(bcc)
    graph_post_action("/me/sendMail", {"message": message})
    return f"메일 발송 완료 → {to} (제목: {subject})"

@mcp.tool()
def reply_email(message_id: str, body: str, reply_all: bool = False) -> str:
    """
    메일 답장
    - message_id: 원본 메일 ID
    - body: 답장 내용
    - reply_all: True면 전체 답장 (기본값: False)
    """
    action = "replyAll" if reply_all else "reply"
    graph_post_action(f"/me/messages/{message_id}/{action}", {"comment": body})
    action_label = "전체 답장" if reply_all else "답장"
    return f"메일 {action_label} 완료 (원본 ID: {message_id})"

@mcp.tool()
def forward_email(message_id: str, to: str, comment: str = "") -> str:
    """
    메일 전달
    - message_id: 원본 메일 ID
    - to: 전달 대상 이메일 (쉼표로 여러 명 가능)
    - comment: 전달 시 추가 코멘트 (선택)
    """
    graph_post_action(f"/me/messages/{message_id}/forward", {
        "comment": comment,
        "toRecipients": _parse_recipients(to),
    })
    return f"메일 전달 완료 → {to} (원본 ID: {message_id})"

@mcp.tool()
def list_mail_folders() -> str:
    """사용 가능한 메일 폴더 목록 조회"""
    data = graph_get("/me/mailFolders", params={"$select": "id,displayName,totalItemCount,unreadItemCount"})
    folders = data.get("value", [])
    result = []
    for f in folders:
        result.append(f"{f['displayName']} (전체: {f['totalItemCount']}, 안읽음: {f['unreadItemCount']}) | ID: {f['id']}")
    return "\n".join(result)

# ═══════════════════════════════════════════
# Teams 파일 (SharePoint 기반)
# ═══════════════════════════════════════════

@mcp.tool()
def list_channel_files(team_id: str, channel_id: str, top: int = 50, skip: int = 0, next_link: str = "") -> str:
    """
    팀 채널에 업로드된 파일 목록 조회
    - team_id: 팀 ID
    - channel_id: 채널 ID
    - top: 가져올 파일 수 (최대 200)
    - skip: 건너뛸 파일 수 (페이지네이션, 기본값 0)
    - next_link: 다음 페이지 링크 (이전 결과에서 제공, 직접 입력 불필요)
    """
    if next_link:
        items = graph_get("", url=next_link)
        drive_id = ""
    else:
        data = graph_get(f"/teams/{team_id}/channels/{channel_id}/filesFolder")
        drive_id = data.get("parentReference", {}).get("driveId", "")
        folder_id = data.get("id", "")
        if not drive_id or not folder_id:
            return "파일 폴더 정보를 가져올 수 없습니다."
        top = min(top, 200)
        params = {
            "$select": "id,name,size,lastModifiedDateTime,webUrl,file,folder",
            "$top": top
        }
        if skip > 0:
            params["$skip"] = skip
        items = graph_get(f"/drives/{drive_id}/items/{folder_id}/children", params=params)
    files = items.get("value", [])
    if not files:
        return "파일이 없습니다."
    # next_link 사용 시 첫 번째 파일의 parentReference에서 driveId 추출
    if not drive_id and files:
        drive_id = files[0].get("parentReference", {}).get("driveId", "")
    result = []
    for i, f in enumerate(files, 1):
        ftype = "📁 폴더" if "folder" in f else "📄 파일"
        size = f.get("size", 0)
        size_str = f"{size / 1024 / 1024:.1f}MB" if size > 1024 * 1024 else f"{size / 1024:.1f}KB"
        result.append(
            f"{i}. {ftype} {f['name']} ({size_str})\n"
            f"   수정: {f.get('lastModifiedDateTime', '')[:19]}\n"
            f"   ID: {f['id']}\n"
            f"   DriveID: {drive_id}"
        )
    return "\n\n".join(result) + _pagination_footer(items, skip, top)

@mcp.tool()
def read_channel_file(drive_id: str, item_id: str) -> str:
    """
    팀 채널의 파일 내용 읽기 (텍스트 기반 파일만 가능)
    - drive_id: list_channel_files에서 확인한 DriveID
    - item_id: list_channel_files에서 확인한 파일 ID
    """
    meta = graph_get(f"/drives/{drive_id}/items/{item_id}")
    name = meta.get("name", "")
    size = meta.get("size", 0)
    download_url = meta.get("@microsoft.graph.downloadUrl", "")

    if size > 5 * 1024 * 1024:
        return f"파일이 너무 큽니다 ({size / 1024 / 1024:.1f}MB). 5MB 이하 파일만 읽을 수 있습니다."

    if not download_url:
        return "다운로드 URL을 가져올 수 없습니다."

    resp = requests.get(download_url)
    _check_response(resp)

    mime = meta.get("file", {}).get("mimeType", "")
    if "image" in mime or "video" in mime or "audio" in mime:
        return f"바이너리 파일입니다 ({mime}). 텍스트 파일만 읽을 수 있습니다."

    # xlsx 파일 처리
    if name.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
        result_parts = [f"파일명: {name}"]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_str = "\t".join(str(c) if c is not None else "" for c in row)
                rows.append(row_str)
                if len(rows) > 200:
                    break
            result_parts.append(f"\n{'─'*50}\n[Sheet: {sheet_name}]\n" + "\n".join(rows))
        wb.close()
        full = "\n".join(result_parts)
        return full[:8000]

    try:
        content = resp.content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            content = resp.content.decode("cp949")
        except UnicodeDecodeError:
            return "파일 인코딩을 판별할 수 없습니다. 바이너리 파일일 수 있습니다."

    return f"파일명: {name}\n{'─'*50}\n{content[:5000]}"

# ═══════════════════════════════════════════
# 인증 관리
# ═══════════════════════════════════════════

@mcp.tool()
def auth_status() -> str:
    """현재 Microsoft 인증 상태 확인"""
    app = _get_app()
    accounts = app.get_accounts()
    if not accounts:
        return (
            "인증 안됨 - 토큰이 없습니다.\n"
            "authenticate 도구를 호출하여 Device Code Flow로 인증하세요."
        )
    account = accounts[0]
    username = account.get("username", "알 수 없음")
    result = app.acquire_token_silent(SCOPES, account=account)
    if result and "access_token" in result:
        save_cache()
        return f"인증됨 - 계정: {username}\n토큰이 유효합니다."
    else:
        return (
            f"토큰 만료 - 계정: {username}\n"
            "authenticate 도구를 호출하여 재인증하세요."
        )

@mcp.tool()
def authenticate() -> str:
    """
    Device Code Flow로 Microsoft 인증 수행.
    반환되는 URL에 접속하여 코드를 입력하면 인증이 완료됩니다.
    """
    app = _get_app()
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0], force_refresh=True)
        if result and "access_token" in result:
            save_cache()
            return f"이미 유효한 토큰이 있습니다. 계정: {accounts[0]['username']}"

    pub_app = _get_pub_app()
    flow = pub_app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        return f"Device Code Flow 시작 실패: {flow.get('error_description', flow)}"

    print(f"[MCP authenticate] 접속: {flow['verification_uri']} / 코드: {flow['user_code']}", flush=True)

    result = pub_app.acquire_token_by_device_flow(flow)
    if "access_token" in result:
        save_cache()
        account = result.get("id_token_claims", {})
        return (
            f"인증 완료!\n"
            f"계정: {account.get('preferred_username', '확인불가')}\n"
            f"토큰 저장 위치: {TOKEN_CACHE_FILE}"
        )
    else:
        return f"인증 실패: {result.get('error_description', result)}"

# ─────────────────────────────────────────
# CLI: 인증 서브커맨드
# ─────────────────────────────────────────

def _parse_auth_args(args):
    """auth 서브커맨드의 --client-id, --client-secret, --tenant-id 파싱"""
    i = 0
    while i < len(args):
        if args[i] == "--client-id" and i + 1 < len(args):
            os.environ["MS_CLIENT_ID"] = args[i + 1]
            i += 2
        elif args[i] == "--client-secret" and i + 1 < len(args):
            os.environ["MS_CLIENT_SECRET"] = args[i + 1]
            i += 2
        elif args[i] == "--tenant-id" and i + 1 < len(args):
            os.environ["MS_TENANT_ID"] = args[i + 1]
            i += 2
        else:
            i += 1

def cmd_auth():
    """Device Code Flow로 토큰 발급 (headless 서버용)"""
    app = _get_app()
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0], force_refresh=True)
        if result and "access_token" in result:
            save_cache()
            print(f"기존 토큰 갱신 성공! 계정: {accounts[0]['username']}")
            return

    pub_app = _get_pub_app()
    flow = pub_app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        print(f"Device Code Flow 시작 실패: {flow.get('error_description', flow)}")
        sys.exit(1)

    print("=" * 60, file=sys.stderr, flush=True)
    print("  Microsoft 토큰 발급 (Device Code Flow)", file=sys.stderr, flush=True)
    print("=" * 60, file=sys.stderr, flush=True)
    print(f"\n  1. 아무 기기에서 접속: {flow['verification_uri']}", file=sys.stderr, flush=True)
    print(f"  2. 코드 입력: {flow['user_code']}", file=sys.stderr, flush=True)
    print(f"\n  만료: {flow.get('expires_in', 900)}초 이내에 완료하세요", file=sys.stderr, flush=True)
    print("=" * 60, file=sys.stderr, flush=True)
    print("\n인증 대기 중...", file=sys.stderr, flush=True)

    result = pub_app.acquire_token_by_device_flow(flow)
    if "access_token" in result:
        save_cache()
        account = result.get("id_token_claims", {})
        print(f"\n토큰 저장 완료! 계정: {account.get('preferred_username', '확인불가')}")
        print(f"저장 위치: {TOKEN_CACHE_FILE}")
    else:
        print(f"\n토큰 획득 실패: {result.get('error_description', result)}")
        sys.exit(1)

def main():
    if len(sys.argv) > 1:
        command = sys.argv[1]
        if command == "auth":
            _parse_auth_args(sys.argv[2:])
            cmd_auth()
        elif command in ("--version", "-v", "version"):
            print(f"ms-teams-mcp {__version__}")
        else:
            print("사용법:")
            print("  ms-teams-mcp                    # MCP 서버 실행")
            print("  ms-teams-mcp auth               # Device Code Flow 인증")
            print("  ms-teams-mcp auth \\")
            print("    --client-id <ID> \\")
            print("    --client-secret <SECRET> \\")
            print("    --tenant-id <TENANT>                 # CLI 인자로 인증")
            print("  ms-teams-mcp --version           # 버전 확인")
            sys.exit(1)
    else:
        print("Microsoft Teams MCP 서버 시작 중...")
        mcp.run()

if __name__ == "__main__":
    main()
